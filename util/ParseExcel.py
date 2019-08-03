from openpyxl import load_workbook

from config.config import testExcelValuePath as DataPath


class ParseExcel(object):
    """解析excel文件"""
    def __init__(self):
        self.data_path = DataPath
        self.wb = load_workbook(self.data_path)

    def get_row_value(self, sheet_name, raw_no):
        """获取某一行的数据"""
        sh = self.wb[sheet_name]
        row_value_list = []
        for y in range(2, sh.max_column + 1):
            value = sh.cell(raw_no, y).value
            row_value_list.append(value)
        return row_value_list

    def get_column_value(self, sheet_name, col_no):
        """获取某一列的数据"""
        sh = self.wb[sheet_name]
        col_value_list = []
        for x in range(2, sh.max_row + 1):
            value = sh.cell(x, col_no).value
            col_value_list.append(value)
        return col_value_list

    def get_cell_value(self, sheet_name, raw_no, col_no):
        """获取某一个单元格的数据"""
        sh = self.wb[sheet_name]
        value = sh.cell(raw_no, col_no).value
        return value

    def write_cell(self, sheet_name, raw_no, col_no, value):
        """向某个单元格写入数据"""
        sh = self.wb[sheet_name]
        sh.cell(raw_no, col_no).value = value
        self.wb.save(self.data_path)


if __name__ == '__main__':
    p = ParseExcel()
    print(p.get_row_value('126account', 2))
    print(p.get_column_value('126account', 3))
    print(p.get_cell_value('126account', 2, 3))
